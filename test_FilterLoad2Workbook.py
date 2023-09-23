import unittest
import MastodonOOP
from openpyxl import load_workbook
import openpyxl
import os
import datetime


class Test_FilterExcel(unittest.TestCase):
    
    def setUp(self):
        self.toot_true = MastodonOOP.Toot (
            account = [{"id": 123, "username": "Marco"}],
            toot_id = True,
            content = 'dog',
            user_id = True,
            hashtags = [{'name': 'dog', 'url': True, 'history': ''}],
            bookmark = True,
            no_replies = True,
            url = True,
            count_replies = True,
            pubdate = "2023-07-14 13:23:05",
            mentions = True,
            media = [{"id": "123","type":"image"}],
            language = 'en',
            poll = True        
        )

        self.toot_true.pubdate = datetime.datetime.strptime(self.toot_true.pubdate, "%Y-%m-%d %H:%M:%S")

        self.toot_false = MastodonOOP.Toot (
            account = '',
            toot_id = '',
            content = 'sun',
            user_id = '',
            hashtags = '',
            bookmark = '',
            no_replies = '',
            url = '',
            count_replies = '',
            pubdate = '',
            mentions = False,
            media = False,
            language = '',
            poll = False        
        )
    
    def tearDown(self):
        pass
    
    def test_Filter(self):
        poll = MastodonOOP.PollTrigger()
        mentions = MastodonOOP.MentionsTrigger()
        media = MastodonOOP.MediaTrigger()

        toot_list_true = [self.toot_true]
        toot_list_false = [self.toot_false]
        trigger_list = [poll, mentions, media]

        filter_true = MastodonOOP.filter_toots(toots = toot_list_true, triggerlist = trigger_list)
        filter_false = MastodonOOP.filter_toots(toots = toot_list_false, triggerlist = trigger_list)

        self.assertIsNotNone(filter_true)
        self.assertEqual(filter_true, toot_list_true)

        self.assertEqual(filter_false, [])
        
        
    def test_Load_to_Workbook(self):
        temp_filename = 'test_objects.xlsx'
        workbook = openpyxl.Workbook()
        workbook.save(temp_filename)
        toot_list = [self.toot_true]

        # hashtag = 'SMS'
        # toot_list = MastodonOOP.load(hashtag)
        
        MastodonOOP.load_to_workbook(toot_list, temp_filename)

        # Aber hier ist das Problem dass wir Username und Pubdate nur als True speichern, das müssten wir ändern
        # Bei der "Live-Daten"-Methode wäre das kein Problem-

        saved_workbook = load_workbook(temp_filename)
        saved_worksheet = saved_workbook.active

        # Check if the written data matches our expectations
        self.assertEqual(saved_worksheet['A2'].value, toot_list[0].account[0]["username"])
        self.assertEqual(saved_worksheet['B2'].value, toot_list[0].pubdate.replace(tzinfo=None))
        self.assertEqual(saved_worksheet['C2'].value, toot_list[0].content)

        saved_workbook.close()

        os.remove(temp_filename)


    
    
if __name__ == '__main__':
    unittest.main()